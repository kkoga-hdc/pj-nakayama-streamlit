import streamlit as st
import pandas as pd
from datetime import timedelta
from openpyxl import load_workbook
from io import BytesIO
from smtplib import SMTP
from email.mime.text import MIMEText
from email.utils import formatdate
from config import *

def set_page():
    """ページの基本設定"""
    st.set_page_config(page_title="勤怠時間", layout="centered")
    st.title("勤怠時間")

def upload_files():
    """Streamlitにファイルアップローダーを設置し、勤怠関連ファイルのアップロード
       対応ファイル形式: CSV ('当期勤怠', '前期勤怠'), Excel ('総労働時間', '時間外労働')"""
    touki = st.file_uploader('当期勤怠ファイルの読み込み', type='csv', key='touki_csv')
    zenki = st.file_uploader('前期勤怠ファイルの読み込み', type='csv', key='zenki_csv')
    # ラジオボタンで使用する選択肢のリスト
    options = ['総労働時間', '残業']
    selected_option = st.radio("設定するファイル種別を選択してください",options, index=0)

    output_file = st.file_uploader(f'{selected_option}ファイルの読み込み', type='xlsx', key='output_xlsx')

    return touki, zenki, selected_option, output_file

def load_csv_data(uploaded_file, usecols):
    """共通関数
        アップロードされたCSVファイルから指定された列を読み込み、データフレームを返却
       ファイルがNoneの場合や読み込みに失敗した場合はNoneを返却"""
    try:
        if uploaded_file is not None:
            df = pd.read_csv(uploaded_file, encoding='shift_jis', header=5, usecols=usecols)
            df = df[df.iloc[:, 0] != '合計']  # 合計行を除外
            # 時間形式(hh:mm形式)を分単位に変換し、新たな列に格納
            df['minutes'] = df.iloc[:, 1].apply(convert_time_to_minutes)
            return df
        else:
            return None
    except Exception as e:
        st.error(f'ファイル読み込みエラー: {e}')
        return None

def convert_time_to_minutes(time_str):
    """
    時間の形式の文字列（例: 'hh:mm'）を受け取り、それを分単位の整数に変換する関数。
    Args:
        time_str (str): コロン ':' で区切られた時間と分の形式の文字列。
    Returns:
        int: 指定された時間を分単位で表した整数。時間データが不適切な場合は0を返す。
    """
    if isinstance(time_str, str) and ':' in time_str:
        hours, minutes = map(int, time_str.split(':'))  # 時間と分を分けて整数に変換
        total_month_minutes = hours * 60 + minutes  # 全体の時間を分に換算
        return total_month_minutes
    else:
        # 非適切なデータは0分として扱う
        return 0

def load_excel_data(uploaded_file):
    """共通関数
        アップロードされたExcelファイルから最初のシートを読み込み、workbookとsheetを返却
       ファイルがNoneの場合や読み込みに失敗した場合はNoneを返却"""
    try:
        if uploaded_file is not None:
            memory = BytesIO(uploaded_file.read())
            workbook = load_workbook(memory)
            sheet = workbook.worksheets[0] # 最初のシートを取得
            return workbook, sheet
        else :
            return None, None
    except Exception as e:
        st.error(f'ファイル読み込みエラー: {e}')
        return None, None

def update_excel_sheet(account_period, sheet, df, selected_month):
    """Excelファイルのシートを更新する共通関数
        指定された月に対応する列を更新し、Excelシートにデータを反映"""
        
    # 1列目(社員CD)をキーに辞書型に変換
    dict_df = df.set_index(df.columns[0]).to_dict(orient='index')
        
    headers = [cell.value for cell in sheet[2]] # ヘッダー情報を取得
    month_col = headers.index(selected_month) + 1     # 更新する列番号を特定
    total_individual_col = headers.index("個人別合計") + 1 
   
   # sheetをループして、data辞書の社員CDと一致する行を検索して更新
    total_month_minutes  = 0
    # total_individual_minutes = 0
    for i, row in enumerate(sheet.iter_rows(min_row=3, min_col=1, max_col=sheet.max_column), start=3):
        employee_cd = row[0].value
        if employee_cd in dict_df:
            # 会社員CDと担当者CDが一致する場合
            # 前期の場合は同一行、当期の場合は一行下
            row_index = i if account_period == "前期" else i + 1
            sheet.cell(row=row_index, column=month_col).value = dict_df[employee_cd][df.columns[1]]
            total_month_minutes += dict_df[employee_cd][df.columns[2]]    # 月別時間累積
            # wk_total_individual_time = (sheet.cell(row=row_index, column=total_individual_col).value).total_seconds()
            # min, sec = divmod(wk_total_individual_time, 60)
            # total_individual_time = dict_df[employee_cd][df.columns[2]] + int(min) # 個人別時間合計に当該別個人別時間を加算
            # sheet.cell(row=row_index, column=total_individual_col).value = format_minutes_to_time(total_individual_time)    # 個人別時間合計
            # total_individual_minutes += total_individual_time

    # # data辞書をループし、シートの社員CDと一致する行を検索して更新
    # for employee_cd, minutes in data.items():
    #     for row in sheet.iter_rows(min_row=3, min_col=1, max_col=1):  # 社員CD列だけ走査
    #         if row[0].value == employee_cd:
    #             # 前期の場合は同一行、当期の場合は一行下
    #             row_index = row[0].row if account_period == "前期" else row[0].row + 1
    #             sheet.cell(row=row_index, column=month_col).value = format_minutes_to_time(minutes)
    #             total_month_minutes += minutes
    #             break  # 社員CDが一致したのでループを抜ける          
            
    if total_month_minutes > 0:        
        formatted_month_time = format_minutes_to_time(total_month_minutes)
        # formatted_individual_time = format_minutes_to_time(total_individual_minutes)
        employee_cd = "999999"
        for row in sheet.iter_rows(min_row=3, min_col=1, max_col=1):
            if row[0].value == employee_cd:
                # 前期の場合は同一行、当期の場合は一行下
                row_index = row[0].row if account_period == "前期" else row[0].row + 1
                sheet.cell(row=row_index, column=month_col).value = formatted_month_time
                # sheet.cell(row=row_index, column=total_individual_col).value = formatted_individual_time
                break  # 合計行が更新されたのでループを抜ける

def format_minutes_to_time(total_month_minutes):
    """累計分を時間の形式（hh:mm）にフォーマットする関数。"""
    hours = total_month_minutes // 60
    minutes = total_month_minutes % 60
    return f"{hours:02}:{minutes:02}"

def download_updated_file(workbook, file_name):
    """ダウンロードボタンを設置する共通関数
        更新したExcelファイルをダウンロード用に準備"""
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    st.download_button(label=f'更新された{file_name}ファイルをダウンロード',
                       data=output,
                       file_name=f'更新された{file_name}.xlsx',
                       mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def send_mail():
    """メール設定"""
    
    # SMTPサーバの設定をsecretsから取得
    smtp_server = st.secrets["smtp"]["server"]
    smtp_port = st.secrets["smtp"]["port"]
    smtp_user = st.secrets["smtp"]["user"]
    smtp_password = st.secrets["smtp"]["password"]

    # メール内容の作成
    send_msg = "勤怠時間設定サービス使用メール\r\n"
    message = MIMEText(send_msg, "plain", "utf-8")
    message["Subject"] = USE_MAIL_SUBJECT
    message["From"] = USE_MAIL_FROM
    message["To"] = ",".join(USE_MAIL_TO)
    message['Date'] = formatdate()
    
    # SMTPサーバに接続してメール送信
    with SMTP(smtp_server, smtp_port) as server:
        server.ehlo()  # EHLOコマンドで挨拶
        server.starttls()  # TLSを使用して通信を暗号化
        server.ehlo()  # 再度EHLO
        server.login(smtp_user, smtp_password)  # SMTPサーバにログイン
        server.send_message(message)  # メール送信

def main():
    set_page()
    selected_month = st.selectbox('月を選択してください', [f'{i}月' for i in range(1, 13)])
    touki, zenki, selected_option, output_file = upload_files()

    if st.button('実行'):

        # 入力チェック
        if not touki and not zenki:
            st.error('当期勤怠ファイル、前期勤怠ファイルを選択してください。')
            return
        if not output_file:
            st.error(f'{selected_option}ファイルを選択してください。')
            return
        
        process_files(touki, zenki, selected_option, output_file, selected_month)


def process_files(touki, zenki,  selected_option, output_file, selected_month):
    """ファイルからデータを読み込み、適切に処理し、ダウンロードを提供します。"""
    if selected_option == "総労働時間":
        # 社員CD, 総労働時間列を読み込む
        df_touki = load_csv_data(touki, [0, 12])
        df_zenki = load_csv_data(zenki, [0, 12])
    elif selected_option == "残業":
        # 社員CD, 残業時間列を読み込む
        df_touki = load_csv_data(touki, [0, 11])
        df_zenki = load_csv_data(zenki, [0, 11])

    workbook, sheet = load_excel_data(output_file)
    update_and_download(workbook, sheet, df_touki, df_zenki, selected_option, selected_month)



def update_and_download(workbook, sheet, df_touki, df_zenki, selected_option, month):
    """データフレームからデータを抽出し、シートを更新してファイルをダウンロードします。"""
    if df_touki is not None:
        update_excel_sheet("当期", sheet, df_touki, month)
    if df_zenki is not None:
        update_excel_sheet("前期", sheet, df_zenki, month)

    download_updated_file(workbook, selected_option)

if __name__ == "__main__":
    main()

    if 'send_mail' not in st.session_state:
        send_mail()
        st.session_state['send_mail'] = 'complete'